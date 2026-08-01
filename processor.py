import json
import re
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
PRODUCTS_PATH = BASE_DIR / "products.json"

COLUMN_ALIASES = {
    "Order consignment id": [
        "Order consignment id",
        "Consignment ID",
        "Consignment",
        "Tracking ID",
    ],
    "Merchant order id": [
        "Merchant order id",
        "Merchant Order ID",
        "Order ID",
        "Merchant ID",
    ],
    "Recipient phone": [
        "Recipient phone",
        "Recipient Phone",
        "Phone",
        "Mobile",
        "Mobile Number",
    ],
    "Order description": [
        "Order description",
        "Order Description",
        "Description",
        "Product",
        "Product Name",
    ],
    "Collectable Amount": [
        "Collectable Amount",
        "COD",
        "COD Amount",
        "Amount",
        "Collectable",
    ],
    "Order created at": [
        "Order created at",
        "Order Created At",
        "Created At",
        "Order Date",
        "Date",
    ],
    "Order status": [
        "Order status",
        "Order Status",
        "Status",
        "Delivery Status",
    ],
}

REQUIRED_COLUMNS = (
    "Recipient phone",
    "Order description",
)


def clean_phone(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    digits = re.sub(r"\D", "", text)

    if digits.startswith("00880"):
        digits = digits[2:]

    if digits.startswith("880") and len(digits) >= 13:
        digits = "0" + digits[3:]

    if len(digits) == 10 and digits.startswith("1"):
        digits = "0" + digits

    return digits


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip().replace('"', "")
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    return re.sub(r"\s+", " ", text)


def normalize_match_text(value: object) -> str:
    text = clean_text(value).casefold()
    text = re.sub(r"[_\-–—/\\|,.;:()\[\]{}]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_money(value: object) -> float:
    if pd.isna(value):
        return 0.0

    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return round(float(cleaned or 0), 2)
    except ValueError:
        return 0.0


def parse_date(value: object) -> str:
    if pd.isna(value) or not str(value).strip():
        return date.today().isoformat()

    parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)

    if pd.isna(parsed):
        return date.today().isoformat()

    return parsed.date().isoformat()


def find_column(df: pd.DataFrame, canonical_name: str) -> str | None:
    normalized = {str(column).strip().casefold(): str(column) for column in df.columns}

    for alias in COLUMN_ALIASES.get(canonical_name, [canonical_name]):
        found = normalized.get(alias.strip().casefold())
        if found:
            return found

    return None


def load_product_rules() -> dict[str, dict[str, list[str]]]:
    if not PRODUCTS_PATH.exists():
        raise FileNotFoundError(f"products.json পাওয়া যায়নি: {PRODUCTS_PATH}")

    with PRODUCTS_PATH.open("r", encoding="utf-8") as file:
        raw_data = json.load(file)

    if not isinstance(raw_data, dict):
        raise ValueError("products.json-এর মূল অংশ object/dictionary হতে হবে।")

    rules: dict[str, dict[str, list[str]]] = {}

    for product_name, config in raw_data.items():
        include_keywords: list[Any] = []
        exclude_keywords: list[Any] = []

        # পুরোনো format:
        # "Hair Oil": ["hair oil", "চুলের তেল"]
        if isinstance(config, list):
            include_keywords = config

        # নতুন format:
        # "Hair Oil": {"keywords": [...], "exclude": [...]}
        elif isinstance(config, dict):
            include_keywords = (
                config.get("keywords")
                or config.get("include")
                or config.get("include_keywords")
                or []
            )
            exclude_keywords = (
                config.get("exclude")
                or config.get("exclude_keywords")
                or []
            )
        else:
            raise ValueError(
                f"{product_name}-এর rule list অথবা object হতে হবে।"
            )

        include = [
            normalize_match_text(keyword)
            for keyword in include_keywords
            if normalize_match_text(keyword)
        ]
        exclude = [
            normalize_match_text(keyword)
            for keyword in exclude_keywords
            if normalize_match_text(keyword)
        ]

        if not include:
            continue

        # দীর্ঘ keyword আগে match করলে সাধারণ keyword-এর ভুল match কমে।
        include.sort(key=len, reverse=True)
        exclude.sort(key=len, reverse=True)

        rules[clean_text(product_name)] = {
            "include": include,
            "exclude": exclude,
        }

    if not rules:
        raise ValueError("products.json-এ কোনো ব্যবহারযোগ্য product rule পাওয়া যায়নি।")

    return rules


def detect_product(
    description: object,
    rules: dict[str, dict[str, list[str]]],
) -> str:
    text = normalize_match_text(description)
    if not text:
        return "Unknown Product"

    best_product = "Unknown Product"
    best_score = 0

    for product, config in rules.items():
        includes = config.get("include", [])
        excludes = config.get("exclude", [])

        if any(keyword in text for keyword in excludes):
            continue

        matched = [keyword for keyword in includes if keyword in text]
        if not matched:
            continue

        # বেশি ও দীর্ঘ keyword match হলে সেই product অগ্রাধিকার পাবে।
        score = sum(len(keyword) for keyword in matched) + (len(matched) * 100)

        if score > best_score:
            best_score = score
            best_product = product

    return best_product



def parse_description_items(
    description: object,
    rules: dict[str, dict[str, list[str]]],
) -> list[dict[str, Any]]:
    """Split courier descriptions and extract product + quantity per item.

    Examples supported:
    - Product x 1 Qty
    - Product x2 Qty
    - Product || Product
    """
    raw = clean_text(description)
    if not raw:
        return []

    parts = [clean_text(part) for part in re.split(r"\s*\|\|\s*", raw) if clean_text(part)]
    items: list[dict[str, Any]] = []

    for part in parts:
        quantity = 1
        match = re.search(r"\bx\s*(\d+)\s*(?:qty|pcs?|pieces?)?\b", part, flags=re.IGNORECASE)
        if match:
            try:
                quantity = max(1, int(match.group(1)))
            except (TypeError, ValueError):
                quantity = 1

        items.append({
            "description": part,
            "product": detect_product(part, rules),
            "quantity": quantity,
        })

    return items


def export_eligibility(
    order_product: str,
    items: list[dict[str, Any]],
) -> tuple[bool, str]:
    """Hair/Pain export contains only one product with quantity exactly 1.

    All rows are still stored in the database and included in COD/dashboard/report
    calculations. Only the automatic Hair/Pain Excel is filtered.
    """
    if order_product not in {"Hair Oil", "Pain Oil"}:
        return True, ""

    if len(items) != 1:
        return False, "Mixed product order"

    quantity = int(items[0].get("quantity") or 1)
    if quantity != 1:
        return False, f"Quantity {quantity} order"

    return True, ""

def read_input_file(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
        last_error: Exception | None = None

        for encoding in encodings:
            try:
                return pd.read_csv(path, encoding=encoding, dtype=str)
            except UnicodeDecodeError as error:
                last_error = error

        if last_error:
            raise last_error

    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype=str)

    raise ValueError("শুধু CSV, XLS অথবা XLSX ফাইল গ্রহণযোগ্য।")


def build_duplicate_key(row: pd.Series) -> str:
    consignment = clean_text(row.get("_consignment", "")).casefold()
    merchant_id = clean_text(row.get("_merchant_id", "")).casefold()
    phone = clean_phone(row.get("_phone", ""))
    product = clean_text(row.get("_product", "")).casefold()
    order_date = clean_text(row.get("_date", ""))

    if consignment:
        return f"consignment:{consignment}"

    if merchant_id:
        return f"merchant:{merchant_id}"

    return f"fallback:{phone}|{product}|{order_date}"


def remove_file_duplicates(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """Return unique rows and duplicate rows using only Consignment/Order ID."""
    if df.empty:
        return df.copy(), []

    seen_consignments: set[str] = set()
    seen_merchants: set[str] = set()
    keep_indices: list[Any] = []
    duplicates: list[dict[str, Any]] = []

    for index, row in df.iterrows():
        consignment = clean_text(row.get("_consignment", "")).casefold()
        merchant_id = clean_text(row.get("_merchant_id", "")).casefold()
        reason = ""
        if consignment and consignment in seen_consignments:
            reason = "Duplicate Consignment ID inside uploaded file"
        elif merchant_id and merchant_id in seen_merchants:
            reason = "Duplicate Merchant Order ID inside uploaded file"

        if reason:
            duplicates.append({
                "consignment": clean_text(row.get("_consignment", "")),
                "merchant_id": clean_text(row.get("_merchant_id", "")),
                "phone": clean_phone(row.get("_phone", "")),
                "product": clean_text(row.get("_product", "")) or "Unknown Product",
                "cod": float(row.get("_cod", 0) or 0),
                "order_date": clean_text(row.get("_date", "")),
                "description": clean_text(row.get("_description", "")),
                "duplicate_reason": reason,
                "original_file": "Same uploaded file",
            })
            continue

        keep_indices.append(index)
        if consignment:
            seen_consignments.add(consignment)
        if merchant_id:
            seen_merchants.add(merchant_id)

    return df.loc[keep_indices].copy(), duplicates


def style_excel(path: Path, product_name: str, order_count: int) -> None:
    workbook = load_workbook(path)
    worksheet = workbook["Filtered Data"]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 26

    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_side = Side(style="thin", color="B4C6E7")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, max_row=worksheet.max_row),
        start=2,
    ):
        worksheet.row_dimensions[row_number].height = 21

        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="center")
            cell.border = border

            if row_number % 2 == 0:
                cell.fill = alternate_fill

    for column_cells in worksheet.columns:
        column_index = column_cells[0].column
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        width = min(max(max_length + 3, 14), 34)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    summary = workbook.create_sheet("Summary")
    summary.sheet_view.showGridLines = False
    summary["A1"] = "Product"
    summary["B1"] = product_name
    summary["A2"] = "Total Orders"
    summary["B2"] = order_count
    summary["A3"] = "Generated Date"
    summary["B3"] = date.today().isoformat()

    for cell in summary["A"]:
        cell.font = Font(bold=True)

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 28

    workbook.save(path)



def product_display_name(row: dict[str, Any]) -> str:
    if not bool(row.get("export_eligible", True)):
        return "Mixed Orders"
    product = clean_text(row.get("product")) or "Unknown Product"
    if product.casefold() == "hair oil":
        text = f"{product} {clean_text(row.get('description'))}"
        if re.search(r"\b200\s*ml\b", text, flags=re.IGNORECASE):
            return "Hair Oil 200ml"
        return "Hair Oil 100ml"
    if product.casefold() == "unknown product":
        return "Unknown Product"
    return product


def expected_cod_from_description(description: object) -> float | None:
    totals = re.findall(
        r"\(Total:\s*([0-9]+(?:\.[0-9]+)?)\s*Tk\)",
        clean_text(description),
        flags=re.IGNORECASE,
    )
    if not totals:
        return None
    return round(sum(float(value) for value in totals), 2)

def create_outputs_from_rows(
    rows: list[dict[str, Any]],
    output_dir: Path,
) -> list[dict[str, Any]]:
    """Create product Excel files only from rows newly inserted in database."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not rows:
        return []

    frame = pd.DataFrame(rows)
    outputs: list[dict[str, Any]] = []

    # সব নতুন order-এর প্রয়োজনীয় XLS তৈরি হয়, কিন্তু delivery rule অনুযায়ী
    # শুধু Hair Oil 100ml ও Pain Oil স্বয়ংক্রিয়ভাবে Telegram-এ পাঠানো হবে।
    if frame.empty:
        return []

    frame["_display_product"] = frame.apply(
        lambda row: product_display_name(row.to_dict()), axis=1
    )

    for product, group in frame.groupby("_display_product", sort=False):
        group = group.copy()
        count = len(group)
        safe_name = re.sub(r'[\/:*?"<>|]+', "-", clean_text(product)).strip()
        safe_name = safe_name or "Unknown Product"

        # আগের কাজের filename logic: Order Date + Product + Order count.
        # Example: Friday 30 July - Hair Oil 200ml - 1 Orders.xlsx
        order_dates = [parse_date(value) for value in group.get("order_date", []) if clean_text(value)]
        unique_dates = sorted(set(order_dates))
        if len(unique_dates) == 1:
            try:
                file_date = pd.to_datetime(unique_dates[0]).strftime("%A %d %B")
            except Exception:
                file_date = unique_dates[0]
        else:
            file_date = date.today().strftime("%A %d %B")

        filename = f"{file_date} - {safe_name} - {count} Orders.xlsx"
        output_path = output_dir / filename

        result = pd.DataFrame(
            {
                "Number": group["phone"],
                "Consignment": group["consignment"],
                "ID": group["merchant_id"],
            }
        )
        result.to_excel(
            output_path,
            index=False,
            sheet_name="Filtered Data",
            engine="openpyxl",
        )
        style_excel(output_path, str(product), count)

        outputs.append(
            {
                "product": str(product),
                "count": count,
                "cod": round(float(pd.to_numeric(group["cod"], errors="coerce").fillna(0).sum()), 2),
                "path": output_path,
                "filename": filename,
                "delivery": "auto" if str(product) in {"Hair Oil 100ml", "Pain Oil"} else "on_demand",
            }
        )

    return outputs



def style_analysis_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_side = Side(style="thin", color="B4C6E7")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in cells:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if row_number % 2 == 0:
                    cell.fill = alternate_fill
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 3, 14), 48)
    workbook.save(path)


def create_analysis_outputs(rows: list[dict[str, Any]], output_dir: Path) -> dict[str, Any]:
    """Create optional quantity, status and COD mismatch XLS reports."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if not rows:
        return {"outputs": [], "quantity_summary": [], "status_summary": {}, "cod_mismatches": 0}

    rules = load_product_rules()
    quantity_counts: dict[tuple[str, str], int] = {}
    status_counts: dict[str, int] = {}
    item_rows: list[dict[str, Any]] = []
    mismatch_rows: list[dict[str, Any]] = []

    for row in rows:
        status = clean_text(row.get("status")) or "Unknown Status"
        status_counts[status] = status_counts.get(status, 0) + 1
        items = parse_description_items(row.get("description"), rules)
        for item in items:
            item_text = clean_text(item.get("description"))
            size_match = re.search(r"\b(100|200)\s*ml\b", item_text, flags=re.I)
            size = f"{size_match.group(1)}ml" if size_match else ""
            product = clean_text(item.get("product")) or "Unknown Product"
            qty = int(item.get("quantity") or 1)
            quantity_counts[(product, size)] = quantity_counts.get((product, size), 0) + qty
            item_rows.append({
                "Order ID": row.get("merchant_id", ""), "Consignment": row.get("consignment", ""),
                "Phone": row.get("phone", ""), "Product": product, "Size": size,
                "Quantity": qty, "Item Description": item_text, "Status": status,
            })
        expected = expected_cod_from_description(row.get("description"))
        actual = round(float(row.get("cod") or 0), 2)
        if expected is not None and abs(expected - actual) >= 0.01:
            mismatch_rows.append({
                "Order ID": row.get("merchant_id", ""), "Consignment": row.get("consignment", ""),
                "Phone": row.get("phone", ""), "Product": row.get("product", ""),
                "Description": row.get("description", ""), "Expected COD": expected,
                "Collectable Amount": actual, "Difference": round(actual - expected, 2), "Status": status,
            })

    quantity_summary = [
        {"Product": product, "Size": size, "Total Quantity": quantity}
        for (product, size), quantity in sorted(quantity_counts.items())
    ]
    outputs: list[dict[str, Any]] = []

    qty_path = output_dir / f"Quantity & Size Report - {len(rows)} Orders.xlsx"
    with pd.ExcelWriter(qty_path, engine="openpyxl") as writer:
        pd.DataFrame(quantity_summary, columns=["Product", "Size", "Total Quantity"]).to_excel(writer, index=False, sheet_name="Summary")
        pd.DataFrame(item_rows).to_excel(writer, index=False, sheet_name="Order Items")
    style_analysis_workbook(qty_path)
    outputs.append({"title": "Quantity & Size Report", "count": len(rows), "path": qty_path, "filename": qty_path.name})

    status_path = output_dir / f"Status-wise Export - {len(rows)} Orders.xlsx"
    status_frame = pd.DataFrame(rows)
    with pd.ExcelWriter(status_path, engine="openpyxl") as writer:
        for status, count in status_counts.items():
            if "status" in status_frame.columns:
                group = status_frame[status_frame["status"].fillna("Unknown Status").astype(str).str.strip().replace("", "Unknown Status") == status]
            else:
                group = status_frame
            group.to_excel(writer, index=False, sheet_name=re.sub(r'[\\/*?:\[\]]+', '-', status)[:31] or "Unknown")
    style_analysis_workbook(status_path)
    outputs.append({"title": "Status-wise Export", "count": len(rows), "path": status_path, "filename": status_path.name})

    if mismatch_rows:
        cod_path = output_dir / f"COD Price Mismatch - {len(mismatch_rows)} Orders.xlsx"
        pd.DataFrame(mismatch_rows).to_excel(cod_path, index=False, sheet_name="COD Mismatch", engine="openpyxl")
        style_analysis_workbook(cod_path)
        outputs.append({"title": "COD Price Mismatch", "count": len(mismatch_rows), "path": cod_path, "filename": cod_path.name})

    return {"outputs": outputs, "quantity_summary": quantity_summary, "status_summary": status_counts, "cod_mismatches": len(mismatch_rows)}

def process_file(
    input_path: Path,
    output_dir: Path,
    source_filename: str,
) -> dict[str, Any]:
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    df = read_input_file(input_path)
    df.columns = [clean_text(column) for column in df.columns]

    resolved_columns = {
        canonical: find_column(df, canonical)
        for canonical in COLUMN_ALIASES
    }

    missing = [
        column
        for column in REQUIRED_COLUMNS
        if not resolved_columns.get(column)
    ]
    if missing:
        raise ValueError(
            "প্রয়োজনীয় কলাম পাওয়া যায়নি: " + ", ".join(missing)
        )

    consignment_column = resolved_columns.get("Order consignment id")
    merchant_column = resolved_columns.get("Merchant order id")
    phone_column = resolved_columns["Recipient phone"]
    description_column = resolved_columns["Order description"]
    cod_column = resolved_columns.get("Collectable Amount")
    date_column = resolved_columns.get("Order created at")
    status_column = resolved_columns.get("Order status")

    if not consignment_column and not merchant_column:
        raise ValueError("Consignment ID অথবা Merchant Order ID কলাম পাওয়া যায়নি।")
    assert phone_column is not None
    assert description_column is not None

    rules = load_product_rules()

    df["_product"] = df[description_column].map(
        lambda value: detect_product(value, rules)
    )
    df["_items"] = df[description_column].map(
        lambda value: parse_description_items(value, rules)
    )
    df["_export_eligible"] = df.apply(
        lambda row: export_eligibility(row["_product"], row["_items"])[0],
        axis=1,
    )
    df["_export_skip_reason"] = df.apply(
        lambda row: export_eligibility(row["_product"], row["_items"])[1],
        axis=1,
    )
    df["_phone"] = df[phone_column].map(clean_phone)
    df["_consignment"] = df[consignment_column].map(clean_text) if consignment_column else ""
    df["_merchant_id"] = df[merchant_column].map(clean_text) if merchant_column else ""
    df["_description"] = df[description_column].map(clean_text)

    if cod_column:
        df["_cod"] = df[cod_column].map(parse_money)
    else:
        df["_cod"] = 0.0

    # Extra business rule (keeps all existing quantity/multi-product rules):
    # Nigella Massage Oil is normally Pain Oil, but COD above 1040 means
    # the order contains additional item(s), so it belongs to Mixed Orders.
    nigella_mixed_mask = (
        df["_product"].astype(str).str.casefold().eq("pain oil")
        & df["_description"].astype(str).str.contains(
            r"\bnigella\b.*\bmassage\s*oil\b",
            case=False,
            regex=True,
            na=False,
        )
        & (pd.to_numeric(df["_cod"], errors="coerce").fillna(0) > 1040)
    )
    df.loc[nigella_mixed_mask, "_export_eligible"] = False
    df.loc[nigella_mixed_mask, "_export_skip_reason"] = (
        "Nigella Massage Oil COD above 1040"
    )

    if date_column:
        df["_date"] = df[date_column].map(parse_date)
    else:
        df["_date"] = date.today().isoformat()

    original_rows = len(df)

    # Consignment এবং Order ID দুটোই খালি হলে row invalid।
    invalid_mask = (df["_consignment"] == "") & (df["_merchant_id"] == "")
    invalid_consignment_count = int(invalid_mask.sum())
    df = df[~invalid_mask].copy()

    df, file_duplicate_rows = remove_file_duplicates(df)
    file_duplicate_count = len(file_duplicate_rows)

    # Excel output is intentionally created later, after database duplicate
    # checking. This guarantees that previously recorded orders never appear
    # in the automatically exported files.
    outputs: list[dict[str, Any]] = []
    db_rows: list[dict[str, Any]] = []

    for _, row in df.iterrows():
        db_rows.append(
            {
                "consignment": row["_consignment"],
                "merchant_id": row["_merchant_id"],
                "phone": row["_phone"],
                "product": row["_product"],
                "cod": float(row["_cod"]),
                "order_date": row["_date"],
                "description": clean_text(row[description_column]),
                "source_file": clean_text(source_filename),
                "export_eligible": bool(row["_export_eligible"]),
                "export_skip_reason": clean_text(row["_export_skip_reason"]),
                "parsed_items": row["_items"],
                "total_quantity": sum(int(item.get("quantity") or 1) for item in row["_items"]) or 1,
                "status": clean_text(row[status_column]) if status_column else "Unknown Status",
            }
        )

    description_counts: dict[str, int] = {}
    quantity_counts: dict[tuple[str, str], int] = {}
    status_counts: dict[str, int] = {}
    cod_mismatches = 0
    for row in db_rows:
        description = clean_text(row.get("description"))
        if description:
            description_counts[description] = description_counts.get(description, 0) + 1
        status = clean_text(row.get("status")) or "Unknown Status"
        status_counts[status] = status_counts.get(status, 0) + 1
        for item in row.get("parsed_items", []):
            product = clean_text(item.get("product")) or "Unknown Product"
            item_text = clean_text(item.get("description"))
            size_match = re.search(r"\b(100|200)\s*ml\b", item_text, flags=re.IGNORECASE)
            size = f"{size_match.group(1)}ml" if size_match else ""
            key = (product, size)
            quantity_counts[key] = quantity_counts.get(key, 0) + int(item.get("quantity") or 1)
        expected = expected_cod_from_description(description)
        if expected is not None and abs(expected - float(row.get("cod") or 0)) >= 0.01:
            cod_mismatches += 1

    quantity_summary = [
        {"Product": product, "Size": size, "Total Quantity": quantity}
        for (product, size), quantity in sorted(quantity_counts.items())
    ]

    return {
        "description_summary": description_counts,
        "quantity_summary": quantity_summary,
        "status_summary": status_counts,
        "cod_mismatches": cod_mismatches,
        "original_rows": original_rows,
        "input_rows": len(df),
        "invalid_consignment_rows": invalid_consignment_count,
        "file_duplicates": file_duplicate_count,
        "file_duplicate_rows": file_duplicate_rows,
        "outputs": outputs,
        "db_rows": db_rows,
    }

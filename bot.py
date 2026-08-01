import asyncio
import html
import logging
import os
import tempfile
import json
import shutil
import uuid
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    BotCommand,
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import (
    all_duplicates,
    all_orders,
    clear_duplicates,
    courier_summary,
    courier_sync_candidates,
    dashboard_summary,
    delete_import,
    duplicate_summary,
    find_order,
    get_import,
    get_orders_by_batch,
    init_db,
    insert_orders_with_results,
    update_courier_status,
    log_import,
    monthly_report,
    recent_imports,
    report,
    save_duplicate_rows,
    today_iso,
    yesterday_iso,
)
from processor import create_analysis_outputs, create_outputs_from_rows, process_file
from google_sheets import append_orders as sheets_append_orders, delete_orders as sheets_delete_orders, replace_all as sheets_replace_all, setup_tabs as sheets_setup_tabs
from pathao_client import PathaoClient, PathaoError, normalize_status

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "20"))
SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
SAVED_EXPORTS_DIR = Path(__file__).resolve().parent / "saved_exports"
SAVED_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

ADMIN_IDS = {
    int(value.strip())
    for value in os.getenv("ADMIN_IDS", "").split(",")
    if value.strip().isdigit()
}

REPORT_MARKERS = {"qty": "Quantity & Size Report", "status": "Status-wise Export", "cod": "COD Price Mismatch"}
PATHAO_AUTO_SYNC = os.getenv("PATHAO_AUTO_SYNC", "false").strip().lower() in {"1", "true", "yes", "on"}
PATHAO_SYNC_MINUTES = max(5, int(os.getenv("PATHAO_SYNC_MINUTES", "15")))
PATHAO_SYNC_LIMIT = max(1, int(os.getenv("PATHAO_SYNC_LIMIT", "300")))

def optional_reports_keyboard(batch_id: str, outputs: Iterable[dict[str, Any]], report_outputs: Iterable[dict[str, Any]] | None = None) -> InlineKeyboardMarkup | None:
    buttons: list[list[InlineKeyboardButton]] = []
    kind_map = {"Hair Oil 200ml": ("hair200", "📦 Hair Oil 200ml"), "Mixed Orders": ("mixed", "📦 Mixed Orders"), "Unknown Product": ("unknown", "📦 Unknown Products")}
    for item in outputs:
        product = str(item.get("product") or "")
        if item.get("delivery") == "on_demand" and product in kind_map:
            kind, label = kind_map[product]
            buttons.append([InlineKeyboardButton(text=f"{label} ({int(item.get('count') or 0)})", callback_data=f"file:{kind}:{batch_id}")])
    for item in report_outputs or []:
        title = str(item.get("title") or "")
        count = int(item.get("count") or 0)
        if title == "Quantity & Size Report":
            buttons.append([InlineKeyboardButton(text="📊 Quantity & Size Report", callback_data=f"rpt:qty:{batch_id}")])
        elif title == "Status-wise Export":
            buttons.append([InlineKeyboardButton(text="🚚 Status-wise Excel", callback_data=f"rpt:status:{batch_id}")])
        elif title == "COD Price Mismatch":
            buttons.append([InlineKeyboardButton(text=f"⚠️ COD Mismatch ({count})", callback_data=f"rpt:cod:{batch_id}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None



def main_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Dashboard"), KeyboardButton(text="📅 আজকের হিসাব")],
            [KeyboardButton(text="📂 Files / Imports"), KeyboardButton(text="📤 Export")],
            [KeyboardButton(text="♻️ Duplicates"), KeyboardButton(text="🔍 Find Order")],
            [KeyboardButton(text="🔄 Pathao Sync"), KeyboardButton(text="↩️ Return Report")],
            [KeyboardButton(text="📈 Reports"), KeyboardButton(text="⚙️ Admin Panel")],
            [KeyboardButton(text="❓ Help")],
        ],
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="নিচের Menu থেকে একটি অপশন বেছে নিন",
    )


def reports_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📅 আজ"), KeyboardButton(text="📅 গতকাল")],
            [KeyboardButton(text="🗓 চলতি মাস"), KeyboardButton(text="📊 সর্বমোট")],
            [KeyboardButton(text="🧴 Hair Oil"), KeyboardButton(text="🧴 Pain Oil")],
            [KeyboardButton(text="📆 নির্দিষ্ট তারিখ"), KeyboardButton(text="📆 নির্দিষ্ট মাস")],
            [KeyboardButton(text="⬅️ মূল Menu")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def admin_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 Import History"), KeyboardButton(text="📂 File Manager")],
            [KeyboardButton(text="📤 Full Export"), KeyboardButton(text="♻️ Duplicate Manager")],
            [KeyboardButton(text="📊 Dashboard"), KeyboardButton(text="❓ Help")],
            [KeyboardButton(text="⬅️ মূল Menu")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def money(value: float | int | None) -> str:
    amount = float(value or 0)
    return f"৳{amount:,.0f}"


def description_summary_lines(rows: Iterable[dict[str, Any]], limit: int = 25) -> list[str]:
    counts = Counter(
        str(row.get("description") or "").strip()
        for row in rows
        if str(row.get("description") or "").strip()
    )
    if not counts:
        return []
    lines = ["", "📦 <b>ফাইলে পাওয়া Product/Description</b>"]
    items = counts.most_common()
    for description, count in items[:limit]:
        lines.append(f"• {html.escape(description)} — <b>{count}টি</b>")
    if len(items) > limit:
        lines.append(f"• আরও {len(items) - limit} ধরনের Description আছে।")
    return lines


def is_admin(message: Message) -> bool:
    if not ADMIN_IDS:
        return True
    return bool(message.from_user and message.from_user.id in ADMIN_IDS)


async def require_admin(message: Message) -> bool:
    if is_admin(message):
        return True

    await message.answer("⛔ এই কমান্ডটি শুধু Admin ব্যবহার করতে পারবেন।")
    return False


def format_report(title: str, rows: Iterable[Any]) -> str:
    rows = list(rows)

    if not rows:
        return f"<b>{html.escape(title)}</b>\n\nকোনো হিসাব পাওয়া যায়নি।"

    total_orders = sum(int(row["orders"] or 0) for row in rows)
    total_cod = sum(float(row["cod"] or 0) for row in rows)

    lines = [f"<b>{html.escape(title)}</b>", ""]

    for row in rows:
        product = html.escape(str(row["product"]))
        orders = int(row["orders"] or 0)
        cod = money(row["cod"])
        lines.append(f"• <b>{product}</b>: {orders} অর্ডার | COD {cod}")

    lines.extend(
        [
            "",
            f"<b>মোট অর্ডার:</b> {total_orders}",
            f"<b>মোট COD:</b> {money(total_cod)}",
        ]
    )

    return "\n".join(lines)


def parse_user_date(raw: str) -> str | None:
    raw = raw.strip()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def parse_year_month(raw: str) -> tuple[int, int] | None:
    raw = raw.strip()

    for fmt in ("%Y-%m", "%m-%Y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.year, parsed.month
        except ValueError:
            continue

    return None


async def send_help(message: Message) -> None:
    await message.answer(
        "📦 <b>Smart Product Accounts Bot V5</b>\n\n"
        "Excel/CSV ফাইল পাঠালে Product শনাক্ত করে হিসাব সংরক্ষণ করবে। "
        "Consignment ID ও Merchant Order ID দিয়ে Duplicate পরীক্ষা হবে।\n\n"
        "নিচের স্থায়ী Menu থেকে প্রয়োজনীয় অপশন চাপুন। ফাইল Import করতে "
        "সরাসরি Excel/CSV ফাইল পাঠান।\n\n"
        "<b>বিশেষ কমান্ড</b>\n"
        "/date 2026-07-28 — নির্দিষ্ট দিনের হিসাব\n"
        "/monthly 2026-07 — নির্দিষ্ট মাসের হিসাব\n"
        "/find CONSIGNMENT — অর্ডার খুঁজুন",
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(),
    )


async def dashboard_cmd(message: Message) -> None:
    data = await asyncio.to_thread(dashboard_summary)

    lines = [
        "📊 <b>Dashboard</b>",
        "",
        "<b>আজ</b>",
        f"• অর্ডার: {data['today_orders']}",
        f"• COD: {money(data['today_cod'])}",
        "",
        "<b>চলতি মাস</b>",
        f"• অর্ডার: {data['month_orders']}",
        f"• COD: {money(data['month_cod'])}",
        "",
        "<b>সর্বমোট</b>",
        f"• অর্ডার: {data['total_orders']}",
        f"• COD: {money(data['total_cod'])}",
    ]

    products = list(data.get("products", []))
    if products:
        lines.extend(["", "<b>Product-wise</b>"])
        for row in products:
            lines.append(
                f"• {html.escape(str(row['product']))}: "
                f"{int(row['orders'] or 0)} অর্ডার | {money(row['cod'])}"
            )

    await message.answer("\n".join(lines), parse_mode=ParseMode.HTML)


async def today_cmd(message: Message) -> None:
    day = today_iso()
    rows = await asyncio.to_thread(report, day, day)
    await message.answer(
        format_report(f"📅 আজকের হিসাব ({day})", rows),
        parse_mode=ParseMode.HTML,
    )


async def yesterday_cmd(message: Message) -> None:
    day = yesterday_iso()
    rows = await asyncio.to_thread(report, day, day)
    await message.answer(
        format_report(f"📅 গতকালের হিসাব ({day})", rows),
        parse_mode=ParseMode.HTML,
    )


async def date_cmd(message: Message) -> None:
    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("ব্যবহার: <code>/date 2026-07-27</code>", parse_mode=ParseMode.HTML)
        return

    parsed = parse_user_date(parts[1])
    if not parsed:
        await message.answer("❌ তারিখ দিন: 2026-07-27 অথবা 27-07-2026")
        return

    rows = await asyncio.to_thread(report, parsed, parsed)
    await message.answer(
        format_report(f"📅 {parsed} এর হিসাব", rows),
        parse_mode=ParseMode.HTML,
    )


async def month_cmd(message: Message) -> None:
    today = date.today()
    start = today.replace(day=1).isoformat()
    end = today.isoformat()

    rows = await asyncio.to_thread(report, start, end)
    await message.answer(
        format_report(f"🗓️ চলতি মাসের হিসাব ({start} — {end})", rows),
        parse_mode=ParseMode.HTML,
    )


async def monthly_cmd(message: Message) -> None:
    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        today = date.today()
        year, month = today.year, today.month
    else:
        parsed = parse_year_month(parts[1])
        if not parsed:
            await message.answer(
                "❌ ব্যবহার: <code>/monthly 2026-07</code>",
                parse_mode=ParseMode.HTML,
            )
            return
        year, month = parsed

    try:
        rows = await asyncio.to_thread(monthly_report, year, month)
    except ValueError:
        await message.answer("❌ মাসটি সঠিক নয়।")
        return

    await message.answer(
        format_report(f"🗓️ মাসিক হিসাব ({year:04d}-{month:02d})", rows),
        parse_mode=ParseMode.HTML,
    )


async def summary_cmd(message: Message) -> None:
    rows = await asyncio.to_thread(report)
    await message.answer(
        format_report("📊 সর্বমোট হিসাব", rows),
        parse_mode=ParseMode.HTML,
    )


async def hair_cmd(message: Message) -> None:
    rows = await asyncio.to_thread(report, None, None, "Hair Oil")
    await message.answer(
        format_report("🧴 Hair Oil-এর মোট হিসাব", rows),
        parse_mode=ParseMode.HTML,
    )


async def pain_cmd(message: Message) -> None:
    rows = await asyncio.to_thread(report, None, None, "Pain Oil")
    await message.answer(
        format_report("🧴 Pain Oil-এর মোট হিসাব", rows),
        parse_mode=ParseMode.HTML,
    )


async def find_cmd(message: Message) -> None:
    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer(
            "ব্যবহার: <code>/find CONSIGNMENT</code> অথবা "
            "<code>/find PHONE</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    value = parts[1].strip()
    row = await asyncio.to_thread(find_order, value)

    if not row:
        await message.answer("❌ অর্ডার পাওয়া যায়নি।")
        return

    await message.answer(
        "🔎 <b>অর্ডার পাওয়া গেছে</b>\n\n"
        f"<b>Product:</b> {html.escape(str(row['product']))}\n"
        f"<b>তারিখ:</b> {html.escape(str(row['order_date']))}\n"
        f"<b>COD:</b> {money(row['cod'])}\n"
        f"<b>Consignment:</b> <code>{html.escape(str(row['consignment'] or ''))}</code>\n"
        f"<b>ID:</b> <code>{html.escape(str(row['merchant_id'] or ''))}</code>\n"
        f"<b>Number:</b> <code>{html.escape(str(row['phone'] or ''))}</code>\n"
        f"<b>Source:</b> {html.escape(str(row['source_file'] or ''))}",
        parse_mode=ParseMode.HTML,
    )


def style_export_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 3, 13),
                42,
            )


def create_export(path: Path) -> int:
    rows = all_orders()
    data = [dict(row) for row in rows]

    columns = [
        "order_date",
        "product",
        "cod",
        "phone",
        "consignment",
        "merchant_id",
        "description",
        "source_file",
        "imported_at",
    ]

    dataframe = pd.DataFrame(data, columns=columns)
    dataframe.rename(
        columns={
            "order_date": "Date",
            "product": "Product",
            "cod": "COD",
            "phone": "Number",
            "consignment": "Consignment",
            "merchant_id": "ID",
            "description": "Description",
            "source_file": "Source File",
            "imported_at": "Imported At",
        },
        inplace=True,
    )

    summary_data = [
        {
            "Product": row["product"],
            "Orders": row["orders"],
            "Total COD": row["cod"],
        }
        for row in report()
    ]

    import_data = [dict(row) for row in recent_imports(100)]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="All Orders")
        pd.DataFrame(
            summary_data,
            columns=["Product", "Orders", "Total COD"],
        ).to_excel(writer, index=False, sheet_name="Summary")

        pd.DataFrame(
            import_data,
            columns=[
                "id",
                "source_file",
                "total_rows",
                "inserted_rows",
                "duplicate_rows",
                "invalid_rows",
                "imported_at",
            ],
        ).to_excel(writer, index=False, sheet_name="Import History")

        style_export_workbook(writer)

    return len(rows)


async def export_cmd(message: Message) -> None:
    if not await require_admin(message):
        return

    status = await message.answer("⏳ রিপোর্ট তৈরি হচ্ছে…")

    try:
        with tempfile.TemporaryDirectory(prefix="bot_export_") as temp_dir:
            output = (
                Path(temp_dir)
                / f"All Accounts - {date.today().isoformat()}.xlsx"
            )

            count = await asyncio.to_thread(create_export, output)

            await message.answer_document(
                FSInputFile(output, filename=output.name),
                caption=(
                    "✅ সব হিসাব Export হয়েছে\n"
                    f"মোট অর্ডার: {count}"
                ),
            )

        await status.delete()
    except Exception:
        logging.exception("Export failed")
        await status.edit_text("❌ রিপোর্ট Export করা যায়নি।")


async def admin_cmd(message: Message) -> None:
    if not await require_admin(message):
        return

    configured = (
        ", ".join(str(admin_id) for admin_id in sorted(ADMIN_IDS))
        if ADMIN_IDS
        else "সব ব্যবহারকারী (ADMIN_IDS সেট করা নেই)"
    )

    await message.answer(
        "🛡️ <b>Admin Panel</b>\n\n"
        f"<b>Admin access:</b> {html.escape(configured)}\n\n"
        "নিচের Admin Menu থেকে অপশন বেছে নিন।",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )


async def imports_cmd(message: Message) -> None:
    if not await require_admin(message):
        return
    rows = await asyncio.to_thread(recent_imports, 10)
    if not rows:
        await message.answer("কোনো Import history পাওয়া যায়নি।")
        return
    await message.answer("📥 <b>সাম্প্রতিক Import History</b>", parse_mode=ParseMode.HTML)
    for row in rows:
        batch_id = str(row["batch_id"] or "")
        text = (
            f"<b>{html.escape(str(row['source_file']))}</b>\n"
            f"• নতুন: {row['inserted_rows']} | Duplicate: {row['duplicate_rows']}\n"
            f"• যোগ COD: {money(row['added_cod'])}\n"
            f"• সময়: {html.escape(str(row['imported_at']))}"
        )
        buttons = []
        if batch_id:
            buttons.append([InlineKeyboardButton(text="🗑 File + হিসাব Delete", callback_data=f"delimp:{batch_id}")])
        markup = InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None
        await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)


def create_duplicate_export(path: Path) -> int:
    rows = [dict(row) for row in all_duplicates()]
    columns = ["duplicate_at", "duplicate_reason", "original_file", "duplicate_file", "consignment", "merchant_id", "phone", "product", "cod", "order_date", "description"]
    frame = pd.DataFrame(rows, columns=columns)
    frame.rename(columns={
        "duplicate_at": "Duplicate Time", "duplicate_reason": "Reason",
        "original_file": "Original File", "duplicate_file": "Duplicate File",
        "consignment": "Consignment", "merchant_id": "Order ID",
        "phone": "Phone", "product": "Product", "cod": "COD",
        "order_date": "Order Date", "description": "Description",
    }, inplace=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Duplicates")
        style_export_workbook(writer)
    return len(rows)


async def duplicates_cmd(message: Message) -> None:
    if not await require_admin(message):
        return
    data = await asyncio.to_thread(duplicate_summary)
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Export Duplicate Excel", callback_data="dupexport")],
        [InlineKeyboardButton(text="🗑 Clear Duplicate History", callback_data="dupclear")],
    ])
    await message.answer(
        "♻️ <b>Duplicate History</b>\n\n"
        f"আজ: {data['today']}\nএই মাস: {data['month']}\n"
        f"সর্বমোট: {data['total']}\nDuplicate COD: {money(data['cod'])}",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def import_delete_request(callback: CallbackQuery) -> None:
    if not callback.from_user or (ADMIN_IDS and callback.from_user.id not in ADMIN_IDS):
        await callback.answer("Admin only", show_alert=True)
        return
    batch_id = (callback.data or "").split(":", 1)[1]
    row = await asyncio.to_thread(get_import, batch_id)
    if not row:
        await callback.answer("Import পাওয়া যায়নি", show_alert=True)
        return
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Confirm Delete", callback_data=f"delok:{batch_id}"),
        InlineKeyboardButton(text="❌ Cancel", callback_data="delcancel"),
    ]])
    await callback.message.answer(
        "⚠️ <b>এই Import Delete করবেন?</b>\n\n"
        f"File: {html.escape(str(row['source_file']))}\n"
        f"Orders: {row['inserted_rows']}\nCOD: {money(row['added_cod'])}\n\n"
        "File, linked orders, COD এবং import history মুছে যাবে।",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )
    await callback.answer()


async def import_delete_confirm(callback: CallbackQuery) -> None:
    if not callback.from_user or (ADMIN_IDS and callback.from_user.id not in ADMIN_IDS):
        await callback.answer("Admin only", show_alert=True)
        return

    batch_id = (callback.data or "").split(":", 1)[1]
    result = await asyncio.to_thread(delete_import, batch_id)
    if not result["deleted"]:
        await callback.answer("Import পাওয়া যায়নি", show_alert=True)
        return

    # SQLite is the master database. Delete never gets blocked by a temporary
    # Google API problem. First try targeted row deletion; if old Sheet rows do
    # not contain Order Key, rebuild every tab from the remaining SQLite rows.
    sheet_message = "Google Sheet Sync: ✅"
    try:
        deleted_rows = result.get("deleted_rows", [])
        sheet_deleted = await asyncio.to_thread(sheets_delete_orders, deleted_rows)
        expected_minimum = len(deleted_rows)
        if deleted_rows and sheet_deleted < expected_minimum:
            remaining_rows = [dict(row) for row in await asyncio.to_thread(all_orders)]
            await asyncio.to_thread(sheets_replace_all, remaining_rows)
            sheet_message = "Google Sheet Sync: ✅ Full Resync"
        else:
            sheet_message = f"Google Sheet Row Delete: {sheet_deleted}"
    except Exception:
        logging.exception("Google Sheets targeted delete failed for batch %s", batch_id)
        try:
            remaining_rows = [dict(row) for row in await asyncio.to_thread(all_orders)]
            await asyncio.to_thread(sheets_replace_all, remaining_rows)
            sheet_message = "Google Sheet Sync: ✅ Full Resync"
        except Exception:
            logging.exception("Google Sheets fallback resync failed for batch %s", batch_id)
            sheet_message = "Google Sheet Sync: ⚠️ Pending — /resync দিন"

    for name in filter(None, str(result.get("export_files", "")).split("|")):
        try:
            path = SAVED_EXPORTS_DIR / Path(name).name
            if path.exists():
                path.unlink()
        except OSError:
            logging.exception("Could not delete saved export")

    await callback.message.edit_text(
        f"✅ Import Delete হয়েছে\n"
        f"Orders: {result['orders']}\n"
        f"COD বাদ: {money(result['cod'])}\n"
        f"{sheet_message}"
    )
    await callback.answer("Deleted")


async def _send_saved_export(callback: CallbackQuery, batch_id: str, marker: str) -> None:
    history = await asyncio.to_thread(get_import, batch_id)
    if not history:
        await callback.answer("Import পাওয়া যায়নি", show_alert=True); return
    names = [Path(name).name for name in str(history["export_files"] or "").split("|") if name]
    selected = next((name for name in names if marker.casefold() in name.casefold()), "")
    path = SAVED_EXPORTS_DIR / selected
    if not selected or not path.exists():
        await callback.answer("File পাওয়া যায়নি বা hosting restart হয়েছে।", show_alert=True); return
    await callback.message.answer_document(FSInputFile(path, filename=selected.split("__", 1)[-1]), caption=f"✅ {marker}")
    await callback.answer("File পাঠানো হয়েছে")

async def optional_file_callback(callback: CallbackQuery) -> None:
    if not callback.from_user or (ADMIN_IDS and callback.from_user.id not in ADMIN_IDS):
        await callback.answer("Admin only", show_alert=True); return
    parts=(callback.data or "").split(":",2)
    markers={"hair200":"Hair Oil 200ml","mixed":"Mixed Orders","unknown":"Unknown Product"}
    if len(parts)!=3 or parts[1] not in markers:
        await callback.answer("File পাওয়া যায়নি", show_alert=True); return
    await _send_saved_export(callback, parts[2], markers[parts[1]])

async def optional_report_callback(callback: CallbackQuery) -> None:
    if not callback.from_user or (ADMIN_IDS and callback.from_user.id not in ADMIN_IDS):
        await callback.answer("Admin only", show_alert=True); return
    parts=(callback.data or "").split(":",2)
    if len(parts)!=3 or parts[1] not in REPORT_MARKERS:
        await callback.answer("Report পাওয়া যায়নি", show_alert=True); return
    await _send_saved_export(callback, parts[2], REPORT_MARKERS[parts[1]])


async def generic_callback(callback: CallbackQuery) -> None:
    data = callback.data or ""
    if data == "delcancel":
        await callback.message.edit_text("❌ Delete বাতিল করা হয়েছে।")
        await callback.answer()
        return
    if data == "dupexport":
        with tempfile.TemporaryDirectory(prefix="dup_export_") as td:
            path = Path(td) / f"Duplicate Orders - {date.today().isoformat()}.xlsx"
            count = await asyncio.to_thread(create_duplicate_export, path)
            if not count:
                await callback.answer("Duplicate history খালি", show_alert=True)
                return
            await callback.message.answer_document(
                FSInputFile(path, filename=path.name),
                caption=f"♻️ Duplicate Orders: {count}",
            )
        await callback.answer()
        return
    if data == "dupclear":
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Clear", callback_data="dupclearok"),
            InlineKeyboardButton(text="❌ Cancel", callback_data="delcancel"),
        ]])
        await callback.message.answer("⚠️ সব Duplicate history মুছে ফেলবেন?", reply_markup=keyboard)
        await callback.answer()
        return
    if data == "dupclearok":
        count = await asyncio.to_thread(clear_duplicates)
        await callback.message.edit_text(f"✅ {count}টি Duplicate history মুছে ফেলা হয়েছে।")
        await callback.answer()


async def handle_document(message: Message, bot: Bot) -> None:
    if not await require_admin(message):
        return

    document = message.document
    if document is None:
        return

    filename = document.file_name or "uploaded_file"
    extension = Path(filename).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        await message.answer("❌ শুধু .xlsx, .xls অথবা .csv ফাইল পাঠান।")
        return

    if (
        document.file_size
        and document.file_size > MAX_FILE_SIZE_MB * 1024 * 1024
    ):
        await message.answer(
            f"❌ ফাইল সর্বোচ্চ {MAX_FILE_SIZE_MB} MB হতে পারবে।"
        )
        return

    status = await message.answer(
        "⏳ Product শনাক্ত, Duplicate পরীক্ষা ও হিসাব সংরক্ষণ করা হচ্ছে…"
    )

    try:
        with tempfile.TemporaryDirectory(prefix="smart_filter_") as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / f"input{extension}"

            telegram_file = await bot.get_file(document.file_id)
            if not telegram_file.file_path:
                raise RuntimeError("Telegram file path পাওয়া যায়নি।")

            await bot.download_file(
                telegram_file.file_path,
                destination=input_path,
            )

            batch_id = uuid.uuid4().hex
            result = await asyncio.to_thread(process_file, input_path, temp_path, filename)
            await asyncio.to_thread(save_duplicate_rows, result.get("file_duplicate_rows", []), batch_id, filename)

            inserted_rows, duplicate_rows = await asyncio.to_thread(
                insert_orders_with_results, result["db_rows"], batch_id
            )
            inserted = len(inserted_rows)
            database_duplicates = len(duplicate_rows)
            sheet_synced = 0
            try:
                sheet_synced = await asyncio.to_thread(sheets_append_orders, inserted_rows)
            except Exception:
                logging.exception("Google Sheets append sync failed for batch %s", batch_id)
            export_skipped_rows = [
                row for row in inserted_rows
                if not bool(row.get("export_eligible", True))
            ]

            # শুধু নতুন এবং Database-এ সফলভাবে যোগ হওয়া Order Export হবে।
            # আগে রেকর্ড হওয়া Duplicate Order কোনো Product Excel-এ থাকবে না।
            result["outputs"] = await asyncio.to_thread(
                create_outputs_from_rows, inserted_rows, temp_path,
            )
            analysis = await asyncio.to_thread(create_analysis_outputs, inserted_rows, temp_path)
            result["analysis_outputs"] = analysis["outputs"]

            file_duplicates = int(result.get("file_duplicates", 0))
            invalid_rows = int(result.get("invalid_consignment_rows", 0))
            total_duplicates = file_duplicates + database_duplicates

            saved_names = []
            for item in result["outputs"] + result["analysis_outputs"]:
                saved_name = f"{batch_id}__{Path(item['filename']).name}"
                shutil.copy2(item["path"], SAVED_EXPORTS_DIR / saved_name)
                saved_names.append(saved_name)

            added_cod = sum(float(row.get("cod") or 0) for row in inserted_rows)
            duplicate_cod = sum(float(row.get("cod") or 0) for row in duplicate_rows) + sum(float(row.get("cod") or 0) for row in result.get("file_duplicate_rows", []))
            await asyncio.to_thread(
                log_import, batch_id, filename,
                int(result.get("original_rows", result["input_rows"])),
                inserted, total_duplicates, invalid_rows,
                added_cod, duplicate_cod, "|".join(saved_names),
            )

            for item in result["outputs"]:
                if item.get("delivery") != "auto":
                    continue
                await message.answer_document(
                    document=FSInputFile(
                        item["path"],
                        filename=item["filename"],
                    ),
                    caption=(
                        f"✅ {item['product']}\n"
                        f"অর্ডার: {item['count']}\n"
                        f"COD: {money(item['cod'])}"
                    ),
                )

            lines = ["✅ <b>Processing সম্পন্ন</b>", ""]

            if result["outputs"]:
                for item in result["outputs"]:
                    lines.append(
                        f"• <b>{html.escape(str(item['product']))}</b>: "
                        f"{item['count']} অর্ডার | {money(item['cod'])}"
                    )

                lines.extend(description_summary_lines(inserted_rows))

                quantity_summary = analysis.get("quantity_summary", [])
                if quantity_summary:
                    lines.extend(["", "📊 <b>Quantity ও Size-wise</b>"])
                    for row in quantity_summary:
                        size = f" {html.escape(str(row.get('Size') or ''))}" if row.get('Size') else ""
                        lines.append(
                            f"• {html.escape(str(row.get('Product') or 'Unknown Product'))}{size}: "
                            f"<b>{int(row.get('Total Quantity') or 0)}টি</b>"
                        )

                status_summary = analysis.get("status_summary", {})
                known_status = {k: v for k, v in status_summary.items() if str(k).casefold() != "unknown status"}
                if known_status:
                    lines.extend(["", "🚚 <b>Status-wise</b>"])
                    for order_status, count in known_status.items():
                        lines.append(f"• {html.escape(str(order_status))}: <b>{int(count)}টি</b>")

                lines.extend(["", f"⚠️ <b>COD Price Mismatch:</b> {int(analysis.get('cod_mismatches', 0))}টি"])
            else:
                if inserted == 0 and total_duplicates > 0:
                    lines.append("সব অর্ডার Duplicate ছিল—কোনো Product Excel তৈরি হয়নি।")
                else:
                    lines.append("কোনো বৈধ নতুন অর্ডার পাওয়া যায়নি।")

            lines.extend(
                [
                    "",
                    f"<b>মূল ফাইলের Row:</b> {result.get('original_rows', 0)}",
                    f"<b>বৈধ Unique Row:</b> {result.get('input_rows', 0)}",
                    f"<b>Database-এ নতুন:</b> {inserted}",
                    f"<b>Google Sheet-এ Sync:</b> {sheet_synced}",
                    f"<b>Mixed / Quantity &gt; 1:</b> {len(export_skipped_rows)}",
                    f"<b>ফাইলের Duplicate:</b> {file_duplicates}",
                    f"<b>আগে থাকা Duplicate:</b> {database_duplicates}",
                    f"<b>Invalid Consignment:</b> {invalid_rows}",
                ]
            )

            report_markup = optional_reports_keyboard(batch_id, result["outputs"], result["analysis_outputs"])
            if report_markup:
                lines.extend(["", "📥 <b>অতিরিক্ত File বা Report চাইলে নিচের Button চাপুন।</b>"])
            await status.edit_text(
                "\n".join(lines), parse_mode=ParseMode.HTML, reply_markup=report_markup,
            )

    except ValueError as exc:
        await status.edit_text(f"❌ {html.escape(str(exc))}", parse_mode=ParseMode.HTML)
    except Exception:
        logging.exception("Failed to process uploaded file")
        await status.edit_text(
            "❌ ফাইলটি প্রসেস করা যায়নি। কলামের নাম, products.json এবং "
            "ফাইল ফরম্যাট পরীক্ষা করুন।"
        )


async def run_pathao_sync() -> dict[str, int | float | str]:
    client = PathaoClient()
    if not client.configured:
        raise PathaoError("Pathao API credentials পাওয়া যায়নি। Railway Variables-এ PATHAO_* values বসান।")
    candidates = await asyncio.to_thread(courier_sync_candidates, PATHAO_SYNC_LIMIT)
    checked = changed = returned = delivered = failed = 0
    return_cod = 0.0
    for row in candidates:
        consignment = str(row.get("consignment") or "").strip()
        if not consignment:
            continue
        checked += 1
        try:
            info = await client.get_order_status(consignment)
            normalized = normalize_status(info.status)
            updated = await asyncio.to_thread(
                update_courier_status,
                consignment,
                normalized,
                info.cod,
                json.dumps(info.raw, ensure_ascii=False),
            )
            if updated and updated.get("status_changed"):
                changed += 1
                if normalized == "RETURNED":
                    returned += 1
                    return_cod += float(updated.get("courier_cod") or updated.get("cod") or 0)
                elif normalized == "DELIVERED":
                    delivered += 1
        except Exception:
            failed += 1
            logging.exception("Pathao sync failed for %s", consignment)
    return {
        "checked": checked,
        "changed": changed,
        "returned": returned,
        "delivered": delivered,
        "failed": failed,
        "return_cod": return_cod,
    }


async def pathao_sync_cmd(message: Message) -> None:
    if not await require_admin(message):
        return
    status = await message.answer("⏳ Pathao status sync হচ্ছে…")
    try:
        result = await run_pathao_sync()
        text = (
            "✅ <b>Pathao Sync সম্পন্ন</b>\n\n"
            f"চেক করা হয়েছে: {result['checked']}\n"
            f"Status পরিবর্তন: {result['changed']}\n"
            f"নতুন Delivered: {result['delivered']}\n"
            f"নতুন Return: {result['returned']}\n"
            f"Return COD minus: {money(result['return_cod'])}\n"
            f"Failed: {result['failed']}"
        )
        await status.edit_text(text, parse_mode=ParseMode.HTML)
    except Exception as exc:
        logging.exception("Pathao manual sync failed")
        await status.edit_text(
            f"❌ Pathao Sync হয়নি: {html.escape(str(exc))}",
            parse_mode=ParseMode.HTML,
        )


async def return_report_cmd(message: Message) -> None:
    data = await asyncio.to_thread(courier_summary)
    lines = [
        "↩️ <b>Pathao Return & Net COD</b>", "",
        f"মোট Parcel: <b>{data['total_parcels']}</b>",
        f"মোট COD: <b>{money(data['total_cod'])}</b>",
        f"Return Parcel: <b>{data['return_parcels']}</b>",
        f"Return COD: <b>{money(data['return_cod'])}</b>",
        f"Net Parcel: <b>{data['net_parcels']}</b>",
        f"Net COD: <b>{money(data['net_cod'])}</b>",
        f"Delivered: <b>{data['delivered_parcels']}</b> | {money(data['delivered_cod'])}",
        f"Pending/In transit: <b>{data['active_parcels']}</b>",
    ]
    products = list(data.get("products", []))
    if products:
        lines.extend(["", "<b>Product-wise Return</b>"])
        for row in products:
            lines.append(
                f"• {html.escape(str(row['product']))}: "
                f"{int(row['return_parcels'] or 0)}টি | {money(row['return_cod'])}"
            )
    await message.answer("\n".join(lines), parse_mode=ParseMode.HTML)


async def pathao_auto_sync_worker() -> None:
    await asyncio.sleep(15)
    while True:
        try:
            if PATHAO_AUTO_SYNC:
                result = await run_pathao_sync()
                logging.info("Pathao auto sync: %s", result)
        except Exception:
            logging.exception("Pathao auto sync cycle failed")
        await asyncio.sleep(PATHAO_SYNC_MINUTES * 60)


async def resync_cmd(message: Message) -> None:
    if not await require_admin(message):
        return
    status = await message.answer("⏳ SQLite থেকে Google Sheets পুনরায় Sync হচ্ছে…")
    try:
        rows = [dict(row) for row in await asyncio.to_thread(all_orders)]
        count = await asyncio.to_thread(sheets_replace_all, rows)
        await status.edit_text(f"✅ Google Sheets Sync সম্পন্ন\nOrders: {count}")
    except Exception as exc:
        logging.exception("Google Sheets full resync failed")
        await status.edit_text(f"❌ Google Sheets Sync হয়নি: {html.escape(str(exc))}", parse_mode=ParseMode.HTML)


async def menu_button_handler(message: Message) -> None:
    text = (message.text or "").strip()

    if text == "📊 Dashboard":
        await dashboard_cmd(message)
    elif text in {"📅 আজকের হিসাব", "📅 আজ"}:
        await today_cmd(message)
    elif text == "📅 গতকাল":
        await yesterday_cmd(message)
    elif text == "🗓 চলতি মাস":
        await month_cmd(message)
    elif text == "📊 সর্বমোট":
        await summary_cmd(message)
    elif text == "🧴 Hair Oil":
        await hair_cmd(message)
    elif text == "🧴 Pain Oil":
        await pain_cmd(message)
    elif text == "🔄 Pathao Sync":
        await pathao_sync_cmd(message)
    elif text == "↩️ Return Report":
        await return_report_cmd(message)
    elif text == "📈 Reports":
        await message.answer("📈 রিপোর্টের ধরন বেছে নিন।", reply_markup=reports_keyboard())
    elif text == "📆 নির্দিষ্ট তারিখ":
        await message.answer("তারিখ লিখুন এভাবে: <code>/date 2026-07-28</code>", parse_mode=ParseMode.HTML)
    elif text == "📆 নির্দিষ্ট মাস":
        await message.answer("মাস লিখুন এভাবে: <code>/monthly 2026-07</code>", parse_mode=ParseMode.HTML)
    elif text in {"📂 Files / Imports", "📥 Import History", "📂 File Manager"}:
        await imports_cmd(message)
    elif text in {"📤 Export", "📤 Full Export"}:
        await export_cmd(message)
    elif text in {"♻️ Duplicates", "♻️ Duplicate Manager"}:
        await duplicates_cmd(message)
    elif text == "🔍 Find Order":
        await message.answer(
            "Consignment বা Merchant Order ID দিয়ে খুঁজুন:\n"
            "<code>/find VALUE</code>",
            parse_mode=ParseMode.HTML,
        )
    elif text == "⚙️ Admin Panel":
        await admin_cmd(message)
    elif text == "❓ Help":
        await send_help(message)
    elif text == "⬅️ মূল Menu":
        await message.answer("🏠 মূল Menu", reply_markup=main_menu_keyboard())


async def setup_bot_commands(bot: Bot) -> None:
    commands = [
        BotCommand(command="start", description="🏠 মূল Menu"),
        BotCommand(command="dashboard", description="📊 Dashboard"),
        BotCommand(command="imports", description="📂 Files ও Import history"),
        BotCommand(command="duplicates", description="♻️ Duplicate history"),
        BotCommand(command="export", description="📤 সব হিসাব Export"),
        BotCommand(command="today", description="📅 আজকের হিসাব"),
        BotCommand(command="yesterday", description="📅 গতকালের হিসাব"),
        BotCommand(command="month", description="🗓 চলতি মাস"),
        BotCommand(command="summary", description="📊 সর্বমোট হিসাব"),
        BotCommand(command="date", description="📆 নির্দিষ্ট তারিখ"),
        BotCommand(command="monthly", description="📆 নির্দিষ্ট মাস"),
        BotCommand(command="find", description="🔍 অর্ডার খুঁজুন"),
        BotCommand(command="admin", description="⚙️ Admin Panel"),
        BotCommand(command="resync", description="🔄 Google Sheets Sync"),
        BotCommand(command="pathao_sync", description="🚚 Pathao status sync"),
        BotCommand(command="returns", description="↩️ Return ও Net COD"),
        BotCommand(command="help", description="❓ সাহায্য"),
    ]
    await bot.set_my_commands(commands)


async def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError(
            "BOT_TOKEN পাওয়া যায়নি। .env ফাইলে token বসান।"
        )

    logging.basicConfig(
        level=logging.INFO,
        format=(
            "%(asctime)s | %(levelname)s | "
            "%(name)s | %(message)s"
        ),
    )

    init_db()
    try:
        await asyncio.to_thread(sheets_setup_tabs)
    except Exception:
        logging.exception("Google Sheets startup setup failed")

    bot = Bot(token=BOT_TOKEN)
    dispatcher = Dispatcher()

    dispatcher.message.register(send_help, CommandStart())
    dispatcher.message.register(send_help, Command("help"))
    dispatcher.message.register(dashboard_cmd, Command("dashboard"))
    dispatcher.message.register(today_cmd, Command("today"))
    dispatcher.message.register(yesterday_cmd, Command("yesterday"))
    dispatcher.message.register(date_cmd, Command("date"))
    dispatcher.message.register(month_cmd, Command("month"))
    dispatcher.message.register(monthly_cmd, Command("monthly"))
    dispatcher.message.register(summary_cmd, Command("summary"))
    dispatcher.message.register(hair_cmd, Command("hair"))
    dispatcher.message.register(pain_cmd, Command("pain"))
    dispatcher.message.register(find_cmd, Command("find"))
    dispatcher.message.register(export_cmd, Command("export"))
    dispatcher.message.register(admin_cmd, Command("admin"))
    dispatcher.message.register(imports_cmd, Command("imports"))
    dispatcher.message.register(duplicates_cmd, Command("duplicates"))
    dispatcher.message.register(resync_cmd, Command("resync"))
    dispatcher.message.register(pathao_sync_cmd, Command("pathao_sync"))
    dispatcher.message.register(return_report_cmd, Command("returns"))
    dispatcher.callback_query.register(import_delete_request, F.data.startswith("delimp:"))
    dispatcher.callback_query.register(import_delete_confirm, F.data.startswith("delok:"))
    dispatcher.callback_query.register(optional_file_callback, F.data.startswith("file:"))
    dispatcher.callback_query.register(optional_report_callback, F.data.startswith("rpt:"))
    dispatcher.callback_query.register(generic_callback, F.data.in_({"delcancel", "dupexport", "dupclear", "dupclearok"}))
    dispatcher.message.register(
        menu_button_handler,
        F.text.in_({
            "📊 Dashboard", "📅 আজকের হিসাব", "📂 Files / Imports",
            "🔄 Pathao Sync", "↩️ Return Report",
            "📤 Export", "♻️ Duplicates", "🔍 Find Order", "📈 Reports",
            "⚙️ Admin Panel", "❓ Help", "📅 আজ", "📅 গতকাল",
            "🗓 চলতি মাস", "📊 সর্বমোট", "🧴 Hair Oil", "🧴 Pain Oil",
            "📆 নির্দিষ্ট তারিখ", "📆 নির্দিষ্ট মাস", "📥 Import History",
            "📂 File Manager", "📤 Full Export", "♻️ Duplicate Manager",
            "⬅️ মূল Menu",
        }),
    )
    dispatcher.message.register(handle_document, F.document)

    await setup_bot_commands(bot)

    logging.info("Smart Product Accounts Bot v2.0 started")
    auto_sync_task = asyncio.create_task(pathao_auto_sync_worker())
    try:
        await dispatcher.start_polling(bot)
    finally:
        auto_sync_task.cancel()


if __name__ == "__main__":
    asyncio.run(main())

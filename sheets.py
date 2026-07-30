from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from threading import Lock
from xml.sax.saxutils import escape

import gspread
from google.auth.transport.requests import AuthorizedSession
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from config import Settings


SCOPES = (
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
)
HEADERS = ["Date", "Category", "Amount"]
MASTER_SHEET_NAME = "All Expenses"
SUMMARY_SHEET_NAME = "Summary"


class GoogleSheetStore:
    def __init__(self, settings: Settings):
        self.settings = settings
        self._worksheet = None
        self._spreadsheet = None
        self._lock = Lock()

    def _credentials(self):
        return Credentials.from_service_account_info(
            self.settings.credentials, scopes=SCOPES
        )

    def _connect(self):
        client = gspread.authorize(self._credentials())
        spreadsheet = client.open_by_key(self.settings.spreadsheet_id)
        self._spreadsheet = spreadsheet
        try:
            return spreadsheet.worksheet(MASTER_SHEET_NAME)
        except gspread.WorksheetNotFound:
            try:
                old_sheet = spreadsheet.worksheet(self.settings.worksheet_name)
                if old_sheet.title != MASTER_SHEET_NAME:
                    old_sheet.update_title(MASTER_SHEET_NAME)
                return old_sheet
            except gspread.WorksheetNotFound:
                return spreadsheet.add_worksheet(
                    title=MASTER_SHEET_NAME, rows=1000, cols=3
                )

    @property
    def worksheet(self):
        if self._worksheet is None:
            self._worksheet = self._connect()
        return self._worksheet

    def ensure_sheet(self) -> None:
        with self._lock:
            first_row = self.worksheet.row_values(1)
            if not first_row:
                self.worksheet.append_row(HEADERS, value_input_option="USER_ENTERED")
            elif first_row[:3] != HEADERS:
                raise ValueError(
                    "Worksheet-এর প্রথম row অবশ্যই Date, Category, Amount হতে হবে"
                )
            self.worksheet.freeze(rows=1)
            self._ensure_derived_sheets()

    def _get_or_create_sheet(
        self, title: str, rows: int = 1000, cols: int = 3
    ):
        try:
            return self._spreadsheet.worksheet(title)
        except gspread.WorksheetNotFound:
            return self._spreadsheet.add_worksheet(
                title=title, rows=rows, cols=cols
            )

    def _ensure_derived_sheets(self) -> None:
        master_ref = f"'{MASTER_SHEET_NAME}'"
        for category in self.settings.categories:
            category_sheet = self._get_or_create_sheet(
                category, rows=1000, cols=2
            )
            category_formula = (
                f'=IFERROR(FILTER({{{master_ref}!A2:A,{master_ref}!C2:C}},'
                f'{master_ref}!B2:B="{category}"),"")'
            )
            category_sheet.update(
                values=[["Date", "Amount"], [category_formula]],
                range_name="A1:B2",
                value_input_option="USER_ENTERED",
            )
            category_sheet.freeze(rows=1)

        summary_sheet = self._get_or_create_sheet(
            SUMMARY_SHEET_NAME, rows=20, cols=2
        )
        summary_values = [["Category", "Total Amount"]]
        for row_number, category in enumerate(
            self.settings.categories, start=2
        ):
            formula = (
                f'=SUMIF({master_ref}!B:B,A{row_number},{master_ref}!C:C)'
            )
            summary_values.append([category, formula])
        total_row = len(self.settings.categories) + 2
        summary_values.append(
            ["সর্বমোট", f"=SUM(B2:B{total_row - 1})"]
        )
        summary_sheet.update(
            values=summary_values,
            range_name=f"A1:B{total_row}",
            value_input_option="USER_ENTERED",
        )
        summary_sheet.freeze(rows=1)

    def now(self) -> datetime:
        return datetime.now(self.settings.timezone)

    def add_expense(self, category: str, amount: float) -> str:
        date_text = self.now().strftime("%Y-%m-%d")
        with self._lock:
            self.worksheet.append_row(
                [date_text, category, amount], value_input_option="USER_ENTERED"
            )
        return date_text

    def _read_rows(self) -> list[dict]:
        with self._lock:
            values = self.worksheet.get_all_values()

        rows = []
        for row_number, row in enumerate(values[1:], start=2):
            if len(row) < 3:
                continue
            try:
                date_value = datetime.strptime(row[0].strip(), "%Y-%m-%d").date()
                amount = float(
                    row[2].replace(",", "").replace("৳", "").strip()
                )
            except (ValueError, TypeError):
                continue
            rows.append(
                {
                    "row_number": row_number,
                    "date": date_value,
                    "category": row[1].strip(),
                    "amount": amount,
                }
            )
        return rows

    def get_today(self) -> list[dict]:
        today = self.now().date()
        return [row for row in self._read_rows() if row["date"] == today]

    def get_month(self) -> list[dict]:
        now = self.now()
        return [
            row
            for row in self._read_rows()
            if row["date"].year == now.year and row["date"].month == now.month
        ]

    def get_all(self) -> list[dict]:
        return self._read_rows()

    def get_recent(self, limit: int = 10) -> list[dict]:
        rows = self._read_rows()
        return list(reversed(rows[-limit:]))

    def get_expense(self, row_number: int) -> dict | None:
        if row_number < 2:
            return None
        return next(
            (
                row
                for row in self._read_rows()
                if row["row_number"] == row_number
            ),
            None,
        )

    def delete_expense(self, row_number: int) -> dict | None:
        if row_number < 2:
            return None
        with self._lock:
            values = self.worksheet.row_values(row_number)
            if len(values) < 3:
                return None
            try:
                date_value = datetime.strptime(
                    values[0].strip(), "%Y-%m-%d"
                ).date()
                amount = float(
                    values[2].replace(",", "").replace("৳", "").strip()
                )
            except (ValueError, TypeError):
                return None
            deleted = {
                "row_number": row_number,
                "date": date_value,
                "category": values[1].strip(),
                "amount": amount,
            }
            self.worksheet.delete_rows(row_number)
            return deleted

    def export_spreadsheet(self, mime_type: str) -> bytes:
        allowed_types = {
            "application/pdf",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        if mime_type not in allowed_types:
            raise ValueError("Unsupported export format")

        session = AuthorizedSession(self._credentials())
        response = session.get(
            "https://www.googleapis.com/drive/v3/files/"
            f"{self.settings.spreadsheet_id}/export",
            params={"mimeType": mime_type},
            timeout=60,
        )
        response.raise_for_status()
        return response.content

    def get_period_rows(self, period: str, offset: int) -> list[dict]:
        if offset < 0:
            raise ValueError("Offset cannot be negative")
        today = self.now().date()

        if period == "week":
            current_monday = today - timedelta(days=today.weekday())
            start = current_monday - timedelta(weeks=offset)
            end = start + timedelta(days=6)
        elif period == "month":
            month_index = today.year * 12 + today.month - 1 - offset
            year, zero_based_month = divmod(month_index, 12)
            month = zero_based_month + 1
            start = date(year, month, 1)
            if month == 12:
                end = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(year, month + 1, 1) - timedelta(days=1)
        elif period == "all":
            return self._read_rows()
        else:
            raise ValueError("Unsupported period")

        return [
            row for row in self._read_rows() if start <= row["date"] <= end
        ]

    @staticmethod
    def _build_xlsx(rows: list[dict], title: str) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Expenses"
        worksheet.append(["Date", "Category", "Amount"])

        for row in rows:
            worksheet.append(
                [
                    row["date"].strftime("%Y-%m-%d"),
                    row["category"],
                    row["amount"],
                ]
            )

        total_row = len(rows) + 2
        worksheet.cell(total_row, 2, "Total")
        worksheet.cell(total_row, 3, f"=SUM(C2:C{total_row - 1})")
        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 16
        worksheet.column_dimensions["B"].width = 32
        worksheet.column_dimensions["C"].width = 18

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        worksheet.cell(total_row, 2).font = Font(bold=True)
        worksheet.cell(total_row, 3).font = Font(bold=True)
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.oddHeader.center.text = title

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _build_pdf(self, rows: list[dict], title: str) -> bytes:
        font_path = (
            Path(__file__).resolve().parent
            / "assets"
            / "NotoSansBengali.ttf"
        )
        if not font_path.exists():
            raise FileNotFoundError("Bengali PDF font is missing")
        font_name = "NotoSansBengali"
        if font_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(
                TTFont(font_name, str(font_path), shapable=True)
            )

        normal = ParagraphStyle(
            "BengaliNormal",
            fontName=font_name,
            fontSize=9,
            leading=13,
            shaping=1,
        )
        heading = ParagraphStyle(
            "BengaliHeading",
            parent=normal,
            fontSize=15,
            leading=20,
            alignment=1,
            spaceAfter=8,
        )
        header = ParagraphStyle(
            "BengaliHeader",
            parent=normal,
            textColor=colors.white,
            alignment=1,
        )

        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=14 * mm,
            leftMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title=title,
        )
        story = [
            Paragraph(escape(title), heading),
            Spacer(1, 3 * mm),
        ]
        table_data = [
            [
                Paragraph("তারিখ", header),
                Paragraph("ক্যাটাগরি", header),
                Paragraph("পরিমাণ", header),
            ]
        ]
        for row in rows:
            table_data.append(
                [
                    Paragraph(row["date"].strftime("%Y-%m-%d"), normal),
                    Paragraph(escape(row["category"]), normal),
                    Paragraph(f"৳{row['amount']:,.2f}".replace(".00", ""), normal),
                ]
            )
        table_data.append(
            [
                "",
                Paragraph("মোট", normal),
                Paragraph(
                    f"৳{sum(row['amount'] for row in rows):,.2f}".replace(
                        ".00", ""
                    ),
                    normal,
                ),
            ]
        )
        table = LongTable(
            table_data,
            colWidths=[38 * mm, 92 * mm, 36 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C9D6")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EAF2F8")),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(table)
        document.build(story)
        return output.getvalue()

    def export_period(
        self, period: str, offset: int, file_format: str, title: str
    ) -> bytes:
        rows = self.get_period_rows(period, offset)
        if file_format == "xlsx":
            return self._build_xlsx(rows, title)
        if file_format == "pdf":
            return self._build_pdf(rows, title)
        raise ValueError("Unsupported export format")
